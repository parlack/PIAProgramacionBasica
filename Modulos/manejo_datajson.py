import json  # Importa el módulo json para trabajar con datos JSON
import pytz  # Importa pytz para trabajar con zonas horarias
import Modulos.Widgets as wg  # Importa el módulo 'Widgets'
import requests  # Importa la biblioteca requests para realizar solicitudes HTTP
import matplotlib.pyplot as plt  # Importa matplotlib para graficar
from datetime import datetime  # Importa datetime para trabajar con fechas y horas
from colorama import Fore, Style, init  # Importa clases para manejar colores en la terminal
from openpyxl import load_workbook, Workbook  # Importa openpyxl para trabajar con archivos de Excel

citty = 'Monterrey'  # Ciudad predeterminada

# Función para realizar una solicitud a la API del clima
def requestAPI(city):
    global citty
    citty = city.replace(' ', '_')  # Reemplaza espacios en blanco con guiones bajos
    ciudadtoapi = city.replace(' ', '%20')  # Reemplaza espacios en blanco con %20 para la URL
    
    # Realiza la solicitud GET a la API del clima
    response = requests.request("GET", f"https://weather.visualcrossing.com/VisualCrossingWebServices/rest/services/timeline/{ciudadtoapi}?unitGroup=metric&include=hours%2Calerts%2Cevents%2Ccurrent%2Cdays&key=UMWHQ8PEMQ54XH8C2SU2RNVPS&contentType=json")

    if response.status_code != 200:
        # Mensaje de error si no se puede obtener una respuesta válida de la API
        print(Fore.RED, 'Unexpected Status code: ', response.status_code, Style.RESET_ALL)
        print('Intente nuevamente:')
        requestAPI(input('Ingrese ciudad a consultar: '))
    else:
        jsonData = json.dumps(response.json(), indent=4)  # Convierte la respuesta JSON en una cadena con formato JSON
        global fecha
        fecha = json.loads(jsonData)['days'][0]['datetime']  # Obtiene la fecha de la respuesta JSON
        # Guarda la respuesta JSON en un archivo
        with open(f'Reportes_Consulta_API/data_{citty}_{fecha}.json', 'w') as archivo:
            archivo.write(jsonData)
        
        if citty == 'Monterrey':
            # Si la ciudad es Monterrey, guarda la respuesta JSON en un archivo diferente
            with open(f'Reportes_Consulta_API/data_{citty}.json', 'w') as archivo:
                archivo.write(jsonData)

# Función para obtener los datos de la API del clima
def data():
    try:
        with open(f'Reportes_Consulta_API/data_{citty}_{fecha}.json', 'r') as f:
            datos = json.load(f)
        return datos
    except:
        with open(f'Reportes_Consulta_API/data_{citty}.json', 'r') as f:
            datos = json.load(f)
        return datos


# Función para obtener el nombre de la ciudad como una cadena
def getcitystr():
    return str(data()['address'])

# Función para obtener la fecha
def fechas(num):
    return str(data()['days'][num-1]['datetime'])

# Función para obtener la fecha actual en un formato específico
def fechaact():
    fecha_actual = datetime.now()
    fechaformato = fecha_actual.strftime("%d-%m-%Y")
    return fechaformato

# Función para obtener la fecha del calendario
def getdatecalendar():
    timezone = data()['timezone']
    tz = pytz.timezone(timezone)
    datafechaact = datetime.now(tz)
    fechaact = str(datafechaact.strftime('%Y-%m-%d'))
    fecha_d_m_aa = wg.getdatecalendar(fechaact, fechas(15))
    fecha_datetime = datetime.strptime(fecha_d_m_aa, "%m/%d/%y")
    fecha_aaaa_mm_dd = fecha_datetime.strftime("%Y-%m-%d")
    return fecha_aaaa_mm_dd

# Función para verificar si los registros de la fecha son recientes
def checardiasfecha():
    timezone = data()['timezone']
    tz = pytz.timezone(timezone)
    datafechaact = datetime.now(tz)
    fechaact = str(datafechaact.strftime('%Y-%m-%d'))
    fecha1_obj = datetime.strptime(fechaact, "%Y-%m-%d")
    fecha2_obj = datetime.strptime(fechas(15), "%Y-%m-%d")
    diferencia = fecha1_obj - fecha2_obj
    
    if diferencia.days <= 0:
        return True
    else:
        return False

# Función para consultar la temperatura actual estimada
def consultTempActEst():
    timezone = data()['timezone']
    tz = pytz.timezone(timezone)
    datafechaact = datetime.now(tz)
    fechaact = str(datafechaact.strftime('%Y-%m-%d'))
    horaact = str(datafechaact.strftime('%H:00:00'))
    temp = 0;
    for dia in data()['days']:
        fecha = dia['datetime']
        if fecha == fechaact:
            for time in dia['hours']:
                hora_del_dia = time['datetime']
                if horaact == hora_del_dia:
                    temp = time['feelslike']
                    return temp

# Función para consultar la temperatura actual
def consultTempAct():
    return data()['currentConditions']['temp']

# Función para obtener las temperaturas mínimas y máximas de un día específico
def temperaturasMinYmaxXdia():
    fechaconsulta = getdatecalendar()
    formateadafecha = datetime.strptime(fechaconsulta, '%Y-%m-%d').strftime('%d/%m/%Y')
    for dia in data()['days']:
        fecha = dia['datetime']
        if fecha == fechaconsulta:
            print(f'Dia: {formateadafecha}')
            print(Fore.GREEN, 'Min:', dia['tempmin'], '\t Max:', dia['tempmax'], Style.RESET_ALL)
            break

# Función para obtener la probabilidad de precipitación para un día específico
def probprecipitacionxdia():
    fechaconsulta = getdatecalendar()
    formateadafecha = datetime.strptime(fechaconsulta, '%Y-%m-%d').strftime('%d/%m/%Y')
    for dia in data()['days']:
        fecha = dia['datetime']
        if fecha == fechaconsulta:
            precip_prob = dia['precipprob']
            print(Fore.GREEN, f'La probabilidad de precipitación del día {formateadafecha} es:', str(precip_prob) + '%', Style.RESET_ALL)
            break
            
# Función para obtener el clima aproximado de las próximas 12 horas
def Climaprox12h():
    timezone = data()['timezone']
    tz = pytz.timezone(timezone)
    datafechaact = datetime.now(tz)

    horaact = int(datafechaact.strftime('%H'))
    cambiodia = 0;
    for i in range(12):
        horafecha = horaact + i
        formatohora = str(str(horafecha).zfill(2))
        horacamparativa = formatohora + ':00:00'
        
        for dia in data()['days']:
            diaa = int(datafechaact.strftime('%d')) + cambiodia
            formatodia = str(str(diaa).zfill(2))
            fechaact = str(datafechaact.strftime(f'%Y-%m-{formatodia}'))
            fecha = dia['datetime']  
            formateadafecha = datetime.strptime(fecha, '%Y-%m-%d').strftime('%d/%m/%Y')
            
            if fecha == fechaact:
                for time in dia['hours']:
                    hora_del_dia = time['datetime']
                    if horacamparativa == hora_del_dia:
                        Temp = time['feelslike']
                        hora = formatohora + ':00'
                        print(Fore.GREEN, f'{formateadafecha},{hora}')
                        print('Temperatura:', Temp, Style.RESET_ALL)
        if horafecha == 23:
            horaact = horafecha - 24 - i
            cambiodia += 1

# Función para calcular estadísticas
def estadisticas(numero):
    try:
        wb = load_workbook(f"Reportes_datos_numéricos/reporte_{fechas(1)}_{getcitystr()}.xlsx")
        ws = wb.active
    except FileNotFoundError:
        wb = Workbook()
        ws = wb.active
        wb.save(f"Reportes_datos_numéricos/reporte_{fechas(1)}_{getcitystr()}.xlsx")
        wb = load_workbook(f"Reportes_datos_numéricos/reporte_{fechas(1)}_{getcitystr()}.xlsx")
        ws = wb.active
        ws['E1'] = 'Indice UV'
        ws['D1'] = 'Radiacion solar'
        ws['C1'] = 'Humedad'
        ws['B1'] = 'Temperaturas'
        ws['A18'] = 'Primera Fecha:'
        ws['A17'] = 'Promedio:'
        ws['A22'] = 'Mas grados:'
        ws['A23'] = 'Menos grados:'
        ws['B21'] = 'Grados'
        ws['C21'] = 'Fecha'
    
    # Realiza diferentes acciones dependiendo del número proporcionado
    match numero:
        case 1:
            temp_sum = 0
            for i in range(15):
                temp = data()['days'][i]['temp']
                ws[f'B{i+2}'] = temp
                temp_sum += temp
            
            promedio = temp_sum / 15
            ws['B17'] = promedio
            ws['B18'] = data()['days'][0]['datetime']
            wb.save(f"Reportes_datos_numéricos/reporte_{fechas(1)}_{getcitystr()}.xlsx")
            return round(promedio, 2)
            
        case 2:
            for i in range(15):
                ws[f'C{i+2}'] = data()['days'][i]['humidity']
            humidity_sum = 0
            for i in range(15):
                humidity = data()['days'][i]['humidity']
                ws[f'C{i+2}'] = humidity
                humidity_sum += humidity
            
            promedio = humidity_sum / 15
            ws['C17'] = promedio
            ws['C18'] = data()['days'][0]['datetime']
            wb.save(f"Reportes_datos_numéricos/reporte_{fechas(1)}_{getcitystr()}.xlsx")
            return round(promedio, 2)
        
        case 3:
            for i in range(15):
                ws[f'D{i+2}'] = data()['days'][i]['solarradiation']
            radsol_sum = 0
            for i in range(15):
                radsol = data()['days'][i]['solarradiation']
                ws[f'D{i+2}'] = radsol
                radsol_sum += radsol
            
            promedio = radsol_sum / 15
            ws['D17'] = promedio
            ws['D18'] = data()['days'][0]['datetime']
            wb.save(f"Reportes_datos_numéricos/reporte_{fechas(1)}_{getcitystr()}.xlsx")
            return round(promedio, 2)
        
        case 4:
            for i in range(15):
                ws[f'E{i+2}'] = data()['days'][i]['uvindex']
            uvindex_sum = 0
            for i in range(15):
                uvindex = data()['days'][i]['uvindex']
                ws[f'E{i+2}'] = uvindex
                uvindex_sum += uvindex
            
            promedio = uvindex_sum / 15
            ws['E17'] = promedio
            ws['E18'] = data()['days'][0]['datetime']
            wb.save(f"Reportes_datos_numéricos/reporte_{fechas(1)}_{getcitystr()}.xlsx")
            return round(promedio, 2)
        
        case 5:
            tempmaxima = -50
            dia = ''
            for day in data()['days']:
                if tempmaxima <= day['tempmax']:
                    tempmaxima = day['tempmax']
                    dia = day['datetime']
            ws['B22'] = tempmaxima
            ws['C22'] = dia
            wb.save(f"Reportes_datos_numéricos/reporte_{fechas(1)}_{getcitystr()}.xlsx")
            return dia, tempmaxima
            
        case 6:
            tempmin = 50
            dia = ''
            for day in data()['days']:
                if tempmin >= day['tempmin']:
                    tempmin = day['tempmin']
                    dia = day['datetime']
            ws['B23'] = tempmin
            ws['C23'] = dia
            wb.save(f"Reportes_datos_numéricos/reporte_{fechas(1)}_{getcitystr()}.xlsx")
            return dia, tempmin

# Función para generar gráficas
def graficas(graph):
    
    # Realiza diferentes acciones dependiendo del número proporcionado
    match graph:
        # Grafica lineal temperatura por hora
        case 1:
            day_hour = []
            temps = []
            for day in data()['days']:
                for hour in day['hours']:
                    temps.append(hour['temp']) 
                    day_hour.append(str(day['datetime']) + '_' + str(hour['datetime']))
            plt.subplot2grid((2, 2), (0, 0), colspan=2, rowspan=2)
            plt.plot(day_hour, temps)
            plt.xticks(rotation=90, color='white')
            plt.title('Temperaturas por hora')
            plt.savefig(f'Graficas/Temp_p_h_{getcitystr()}_{fechaact()}.png')
            plt.show()
        # Grafica de barra, temperaturas por dia
        case 2:
            fechas = []
            minimos = []
            maximos = []
            
            for day in data()['days']:
                fecha = datetime.strptime(day['datetime'], '%Y-%m-%d')
                dia_mes = fecha.strftime('%d-%m')
                fechas.append(str(dia_mes))
                minimos.append(day['tempmin']) 
                maximos.append(day['tempmax']) 
    
            centros = range(len(fechas))
            alturas = [maximos[i] - minimos[i] for i in range(len(fechas))]
            plt.bar(centros, alturas, bottom=minimos, color='blue', alpha=0.7, align='center', edgecolor='black')
            plt.xlabel('Categorías')
            plt.ylabel('Valores')
            plt.title('Temperaturas maximas y minimas')
            plt.xticks(centros, fechas, rotation=90)
            plt.savefig(f'Graficas/Temp_myn_{getcitystr()}_{fechaact()}.png')
            plt.show()
        # Grafica lineal niveles de humedad
        case 3:
            diafecha = []
            humedad = []
            for day in data()['days']:
                fecha = datetime.strptime(day['datetime'], '%Y-%m-%d')
                dia_mes = fecha.strftime('%d-%m')
                humedad.append(day['humidity']) 
                diafecha.append(str(dia_mes))
                
            
            plt.subplot2grid((2, 2), (0, 0), colspan=2, rowspan=2)
            plt.plot(diafecha, humedad)
            plt.xticks(rotation=90, color='black', fontsize=6,)
            plt.title('Grafica niveles de humedad')
            plt.savefig(f'Graficas/niv_humed_{getcitystr()}_{fechaact()}.png')
            plt.show()
        # Grafica lineal probabilidad de precipitacion
        case 4:
            
            day_hour = []
            precipprob = []
            for day in data()['days']:
                fecha = datetime.strptime(day['datetime'], '%Y-%m-%d')
                dia_mes = fecha.strftime('%d-%m')
                precipprob.append(day['precipprob']) 
                day_hour.append(str(dia_mes))
                
            
            plt.subplot2grid((2, 2), (0, 0), colspan=2, rowspan=2)
            plt.plot(day_hour, precipprob)
            plt.xticks(rotation=90, color='black', fontsize=6)
            plt.title('Grafica probabilidad de precipitaciones')
            plt.savefig(f'Graficas/Prob_prec_{getcitystr()}_{fechaact()}.png')
            plt.show()
        # Grafica lineal indice UV
        case 5:
            day_hour = []
            indiceuv = []
            for day in data()['days']:
                fecha = datetime.strptime(day['datetime'], '%Y-%m-%d')
                dia_mes = fecha.strftime('%d-%m')
                indiceuv.append(day['uvindex']) 
                day_hour.append(str(dia_mes))
            
            plt.subplot2grid((2, 2), (0, 0), colspan=2, rowspan=2)
            plt.plot(day_hour, indiceuv)
            plt.xticks(rotation=90, color='black', fontsize=6)
            plt.title('Grafica Indice UV')
            plt.savefig(f'Graficas/indc_uv_{getcitystr()}_{fechaact()}.png')
            plt.show()


