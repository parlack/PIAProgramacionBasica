import json 
import pytz
import Modulos.Widgets as wg
import requests
import json
from datetime import datetime
from colorama import Fore, Style, init
from openpyxl import load_workbook
from openpyxl import Workbook

def requestAPI(city):
    ciudadtoapi = city.replace(' ','%20')
    response = requests.request("GET", f"https://weather.visualcrossing.com/VisualCrossingWebServices/rest/services/timeline/{ciudadtoapi}?unitGroup=metric&include=hours%2Calerts%2Cevents%2Ccurrent%2Cdays&key=UMWHQ8PEMQ54XH8C2SU2RNVPS&contentType=json")

    if response.status_code!=200:
        print(Fore.RED,'Unexpected Status code: ', response.status_code,Style.RESET_ALL)  
        print('Intente nuevamente:')
        requestAPI(input('Ingrese ciudad a consultar: '))
    else:
      jsonData =  json.dumps(response.json(),indent=4)
      with open('Reportes_Consulta_API/Datos_completos_clima.json', 'w') as archivo:
        archivo.write(jsonData)


def data():
    with open('Reportes_Consulta_API/Datos_completos_clima.json', 'r') as f:
        datos = json.load(f)
    return datos

def getcity():
    return data()['resolvedAddress']

def getcitystr():
    return str(data()['address']) 

def fechas(num):
    return str(data()['days'][num-1]['datetime'])
        
def getdatecalendar():
    timezone = data()['timezone']
    tz = pytz.timezone(timezone)
    datafechaact = datetime.now(tz)
    fechaact=str(datafechaact.strftime('%Y-%m-%d'))
    fecha_d_m_aa  = wg.getdatecalendar(fechaact,fechas(15))
    fecha_datetime = datetime.strptime(fecha_d_m_aa, "%m/%d/%y")
    fecha_aaaa_mm_dd = fecha_datetime.strftime("%Y-%m-%d")
    return fecha_aaaa_mm_dd

def checardiasfecha():
    timezone = data()['timezone']
    tz = pytz.timezone(timezone)
    datafechaact = datetime.now(tz)
    fechaact=str(datafechaact.strftime('%Y-%m-%d'))
    fecha1_obj = datetime.strptime(fechaact, "%Y-%m-%d")
    fecha2_obj = datetime.strptime(fechas(15), "%Y-%m-%d")
    diferencia = fecha1_obj - fecha2_obj
    
    if diferencia.days<=0:
        return True
    else:
        return False
    
def consultTempActEst():
    timezone = data()['timezone']
    tz = pytz.timezone(timezone)
    datafechaact = datetime.now(tz)
    fechaact=str(datafechaact.strftime('%Y-%m-%d'))
    horaact=str(datafechaact.strftime('%H:00:00'))
    temp=0;
    for dia in data()['days']:
        fecha = dia['datetime']
        if fecha==fechaact:
            for time in dia['hours']:
                hora_del_dia = time['datetime']
                if horaact==hora_del_dia:
                    temp= time['feelslike']
                    return temp
                
def consultTempAct():

    return data()['currentConditions']['temp']

def temperaturasMinYmaxXdia():
    fechaconsulta=getdatecalendar()
    for dia in data()['days']:
        fecha = dia['datetime']
        if fecha==fechaconsulta:
            print(f'Dia: {fechaconsulta}')
            print('Min:',dia['tempmin'],'\t Max:',dia['tempmax'])
            break

def probprecipitacionxdia():
    fechaconsulta=getdatecalendar()

    for dia in data()['days']:
        fecha = dia['datetime']
        if fecha==fechaconsulta:
            precip_prob = dia['precipprob']
            print(f'La probabilidad de precipitacion del dia {fechaconsulta} es:',str(precip_prob)+'%')
            
def Climaprox12h():
    timezone = data()['timezone']
    tz = pytz.timezone(timezone)
    datafechaact = datetime.now(tz)

    horaact=int(datafechaact.strftime('%H'))
    cambiodia=0;
    for i in range(12):
        horafecha=horaact+i
        formatohora=str(str(horafecha).zfill(2))
        horacamparativa=formatohora+':00:00'

        if horafecha==23:
            horaact=horafecha-24-i
            cambiodia+=1
            
        for dia in data()['days']:
            diaa=int(datafechaact.strftime('%d'))+cambiodia
            formatodia=str(str(diaa).zfill(2))

            fechaact=str(datafechaact.strftime(f'%Y-%m-{formatodia}'))
            fecha = dia['datetime']
            if fecha==fechaact:
                
                for time in dia['hours']:
                    hora_del_dia = time['datetime']

                    if horacamparativa==hora_del_dia:
                        Temp= time['feelslike']
                        hora=formatohora+':00'
                        print(Fore.GREEN,f'{fecha},{hora}')
                        print('Temperatura:',Temp,Style.RESET_ALL)
       



    
def PromTempProxDias():
    try:
        wb = load_workbook(f"Reportes_datos_numéricos/reporte_{fechas(1)}_{getcitystr()}.xlsx")
        ws = wb.active

    except FileNotFoundError:
        wb = Workbook()
        ws = wb.active
        wb.save(f"Reportes_datos_numéricos/reporte_{fechas(1)}_{getcitystr()}.xlsx")
        wb = load_workbook(f"Reportes_datos_numéricos/reporte_{fechas(1)}_{getcitystr()}.xlsx")
        ws = wb.active
        ws['D1'] = 'Radiacion solar'
        ws['C1'] = 'Humedad'
        ws['B1'] = 'Temperaturas'
        ws['A18'] = 'Primera Fecha:'
        ws['A17'] = 'Promedio:'
        ws['A22'] = 'Mas grados:'
        ws['A23'] = 'Menos grados:'
        ws['B21'] = 'Grados'
        ws['C21'] = 'Fecha'

    temp_sum = 0
    for i in range(15):
        print(data()['days'][i]['temp'])
        temp = data()['days'][i]['temp']
        ws[f'B{i+2}'] = temp
        temp_sum += temp
    
    promedio = temp_sum / 15
    ws['B17'] = promedio
    
    ws['B18'] = data()['days'][0]['datetime']
    
    wb.save(f"Reportes_datos_numéricos/reporte_{fechas(1)}_{getcitystr()}.xlsx")
    return round(promedio, 2)

def PromHumedadProxDias():
    try:
        wb = load_workbook(f"Reportes_datos_numéricos/reporte_{fechas(1)}_{getcitystr()}.xlsx")
        ws = wb.active

    except FileNotFoundError:
        wb = Workbook()
        ws = wb.active
        wb.save(f"Reportes_datos_numéricos/reporte_{fechas(1)}_{getcitystr()}.xlsx")
        wb = load_workbook(f"Reportes_datos_numéricos/reporte_{fechas(1)}_{getcitystr()}.xlsx")
        ws = wb.active
        ws['D1'] = 'Radiacion solar'
        ws['C1'] = 'Humedad'
        ws['B1'] = 'Temperaturas'
        ws['A18'] = 'Primera Fecha:'
        ws['A17'] = 'Promedio:'
        ws['A22'] = 'Mas grados:'
        ws['A23'] = 'Menos grados:'
        ws['B21'] = 'Grados'
        ws['C21'] = 'Fecha'
        
    for i in range(15):
        print(data()['days'][i]['humidity'])
        ws[f'C{i+2}']=data()['days'][i]['humidity']
    humidity_sum = 0
    for i in range(15):
        humidity = data()['days'][i]['humidity']
        ws[f'C{i+2}'] = humidity
        humidity_sum += humidity
    
    promedio = humidity_sum / 15
    ws['C17'] = promedio
    ws['C18'] = data()['days'][0]['datetime']

    wb.save(f"Reportes_datos_numéricos/reporte_{fechas(1)}_{getcitystr()}.xlsx")
    return round(promedio, 2)

def PromRadiacionSolar():
    try:
        wb = load_workbook(f"Reportes_datos_numéricos/reporte_{fechas(1)}_{getcitystr()}.xlsx")
        ws = wb.active

    except FileNotFoundError:
        wb = Workbook()
        ws = wb.active
        wb.save(f"Reportes_datos_numéricos/reporte_{fechas(1)}_{getcitystr()}.xlsx")
        wb = load_workbook(f"Reportes_datos_numéricos/reporte_{fechas(1)}_{getcitystr()}.xlsx")
        ws = wb.active
        ws['D1'] = 'Radiacion solar'
        ws['C1'] = 'Humedad'
        ws['B1'] = 'Temperaturas'
        ws['A18'] = 'Primera Fecha:'
        ws['A17'] = 'Promedio:'
        ws['A22'] = 'Mas grados:'
        ws['A23'] = 'Menos grados:'
        ws['B21'] = 'Grados'
        ws['C21'] = 'Fecha'
        
    for i in range(15):
        print(data()['days'][i]['solarradiation'])
        ws[f'D{i+2}']=data()['days'][i]['solarradiation']
    radsol_sum = 0
    for i in range(15):
        radsol = data()['days'][i]['solarradiation']
        ws[f'D{i+2}'] = radsol
        radsol_sum += radsol
    
    promedio = radsol_sum / 15
    ws['D17'] = promedio
    ws['D18'] = data()['days'][0]['datetime']

    wb.save(f"Reportes_datos_numéricos/reporte_{fechas(1)}_{getcitystr()}.xlsx")
    return round(promedio, 2)

def PromIndiceuv():
    try:
        wb = load_workbook(f"Reportes_datos_numéricos/reporte_{fechas(1)}_{getcitystr()}.xlsx")
        ws = wb.active

    except FileNotFoundError:
        wb = Workbook()
        ws = wb.active
        wb.save(f"Reportes_datos_numéricos/reporte_{fechas(1)}_{getcitystr()}.xlsx")
        wb = load_workbook(f"Reportes_datos_numéricos/reporte_{fechas(1)}_{getcitystr()}.xlsx")
        ws = wb.active
        ws['D1'] = 'Radiacion solar'
        ws['C1'] = 'Humedad'
        ws['B1'] = 'Temperaturas'
        ws['A18'] = 'Primera Fecha:'
        ws['A17'] = 'Promedio:'
        ws['A22'] = 'Mas grados:'
        ws['A23'] = 'Menos grados:'
        ws['B21'] = 'Grados'
        ws['C21'] = 'Fecha'
        
    for i in range(15):
        print(data()['days'][i]['uvindex'])
        ws[f'E{i+2}']=data()['days'][i]['uvindex']
    uvindex_sum = 0
    for i in range(15):
        uvindex = data()['days'][i]['uvindex']
        ws[f'E{i+2}'] = uvindex
        uvindex_sum += uvindex
    
    promedio = uvindex_sum / 15
    ws['E17'] = promedio
    ws['E18'] = data()['days'][0]['datetime']

    wb.save(f"Reportes_datos_numéricos/reporte_{fechas(1)}_{getcitystr()}.xlsx")
    return round(promedio, 2)

def diaMayorTemperatura():
    try:
        wb = load_workbook(f"Reportes_datos_numéricos/reporte_{fechas(1)}_{getcitystr()}.xlsx")
        ws = wb.active

    except FileNotFoundError:
        wb = Workbook()
        ws = wb.active
        wb.save(f"Reportes_datos_numéricos/reporte_{fechas(1)}_{getcitystr()}.xlsx")
        wb = load_workbook(f"Reportes_datos_numéricos/reporte_{fechas(1)}_{getcitystr()}.xlsx")
        ws = wb.active
        ws['D1'] = 'Radiacion solar'
        ws['C1'] = 'Humedad'
        ws['B1'] = 'Temperaturas'
        ws['A18'] = 'Primera Fecha:'
        ws['A17'] = 'Promedio:'
        ws['A22'] = 'Mas grados:'
        ws['A23'] = 'Menos grados:'
        ws['B21'] = 'Grados'
        ws['C21'] = 'Fecha'
        
    tempmaxima=-50
    dia=''
    for day in data()['days']:
        if tempmaxima <= day['tempmax']:
            tempmaxima=day['tempmax']
            dia=day['datetime']
    ws['B22'] = tempmaxima
    ws['C22'] = dia
    wb.save(f"Reportes_datos_numéricos/reporte_{fechas(1)}_{getcitystr()}.xlsx")

    return dia,tempmaxima

def diamenorTemperatura():
    try:
        wb = load_workbook(f"Reportes_datos_numéricos/reporte_{fechas(1)}_{getcitystr()}.xlsx")
        ws = wb.active

    except FileNotFoundError:
        wb = Workbook()
        ws = wb.active
        wb.save(f"Reportes_datos_numéricos/reporte_{fechas(1)}_{getcitystr()}.xlsx")
        wb = load_workbook(f"Reportes_datos_numéricos/reporte_{fechas(1)}_{getcitystr()}.xlsx")
        ws = wb.active
        ws['D1'] = 'Radiacion solar'
        ws['C1'] = 'Humedad'
        ws['B1'] = 'Temperaturas'
        ws['A18'] = 'Primera Fecha:'
        ws['A17'] = 'Promedio:'
        ws['A22'] = 'Mas grados:'
        ws['A23'] = 'Menos grados:'
        ws['B21'] = 'Grados'
        ws['C21'] = 'Fecha'
         
    tempmin=50
    dia=''
    for day in data()['days']:
        if tempmin >= day['tempmin']:
            tempmin=day['tempmin']
            dia=day['datetime']
    ws['B23'] = tempmin
    ws['C23'] = dia
    wb.save(f"Reportes_datos_numéricos/reporte_{fechas(1)}_{getcitystr()}.xlsx")

    return dia,tempmin


    
